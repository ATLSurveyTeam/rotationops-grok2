"""Update Employee Master inside the scheduler workbook without opening Excel."""

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

CORE_FIELDS = [
    "Employee ID",
    "Employee Name",
    "Shift",
    "Days Off",
    "Lead",
    "Mentor",
    "Active",
    "Lunch Preference",
    "Information Desk",
    "Main Inside",
    "Survey",
    "Divest",
    "LOA",
]

SHIFTS = ["3:45 AM", "5:45 AM", "12:15 PM", "2:00 PM", "9:15 PM"]
YN = ["Y", "N"]


def _norm(val) -> str:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    return str(val).strip()


def next_employee_id(employees: pd.DataFrame) -> str:
    nums = []
    for raw in employees.get("Employee ID", pd.Series(dtype=str)).fillna(""):
        text = _norm(raw).upper().replace("EMP-", "")
        if text.isdigit():
            nums.append(int(text))
    nxt = (max(nums) + 1) if nums else 1
    return f"EMP-{nxt:04d}"


def save_employee_master(workbook_path: str | Path, employees: pd.DataFrame) -> None:
    """Write Employee Master rows back, keeping every other sheet intact."""
    path = Path(workbook_path)
    wb = load_workbook(path)
    if "Employee Master" not in wb.sheetnames:
        raise ValueError("Employee Master sheet is missing.")
    ws = wb["Employee Master"]
    headers = [_norm(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
    header_map = {h: i + 1 for i, h in enumerate(headers) if h}

    # Clear old data rows but keep header
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    for r_i, row in employees.iterrows():
        excel_row = ws.max_row + 1
        for col_name, col_idx in header_map.items():
            if col_name in employees.columns:
                val = row.get(col_name, "")
                if pd.isna(val):
                    val = ""
                ws.cell(excel_row, col_idx, _norm(val) if not isinstance(val, (int, float)) else val)

    wb.save(path)


def upsert_employee(employees: pd.DataFrame, fields: dict) -> pd.DataFrame:
    out = employees.copy()
    eid = _norm(fields.get("Employee ID"))
    if not eid:
        eid = next_employee_id(out)
        fields["Employee ID"] = eid
    mask = out["Employee ID"].astype(str).str.strip() == eid
    if mask.any():
        idx = out.index[mask][0]
        for key, val in fields.items():
            if key not in out.columns:
                out[key] = ""
            out.at[idx, key] = val
    else:
        for key in fields:
            if key not in out.columns:
                out[key] = ""
        out = pd.concat([out, pd.DataFrame([fields])], ignore_index=True)
    return out
